from docxtpl import DocxTemplate
from openpyxl import load_workbook


def open_excel_worksheet_by_filename(excel_path):
    excel_path = excel_path.replace('/', '\\')
    wb = load_workbook(excel_path)
    ws = wb.active
    return ws


def open_model(model_path):
    assert model_path != '', 'Empty model path!'
    assert model_path[-5:] == '.docx', 'Model error!'

    model_path = model_path.replace('/', '\\')
    word = DocxTemplate(model_path)
    return word


def date_process(Date, add_day, f_str):
    year, month, day = Date.split('.')
    n_day = int(day) + add_day
    n_month = int(month)
    n_year = int(year)
    if n_day > 28 and n_month == 2 and not ((n_year % 4 == 0 and n_year % 100 != 0) or n_year % 400 == 0):
        n_month += 1
        n_day -= 28
    elif n_day > 29 and n_month == 2 and ((n_year % 4 == 0 and n_year % 100 != 0) or n_year % 400 == 0):
        n_month += 1
        n_day -= 29
    elif n_day > 30 and (n_month == 4 or n_month == 6 or n_month == 9 or n_month == 11):
        n_month += 1
        n_day -= 30
    elif n_day > 31:
        if n_month == 12:
            n_year += 1
            n_month = 1
            n_day -= 31
        else:
            n_month += 1
            n_day -= 31
    n_month = '%02d' % n_month
    n_day = '%02d' % n_day
    if isinstance(f_str, str):
        result = str(n_year) + f_str + str(n_month) + f_str + str(n_day)
    elif isinstance(f_str, list) and len(f_str) == 3:
        result = str(n_year) + f_str[0] + str(n_month) + f_str[1] + str(n_day) + f_str[2]
    return result


def write_information(word, ws, number):
    Place = ws['m' + number].value
    Number = ws['p' + number].value
    Name = ws['b' + number].value
    S_M = ws['c' + number].value
    Supplier = ws['g' + number].value
    O_N = ws['i' + number].value
    G_N = ws['e' + number].value
    U = ws['d' + number].value
    E_N = ws['f' + number].value
    Standard = ws['o' + number].value
    Rate = float(E_N) / float(G_N) * 100
    Rate = '%.3f' % Rate
    Rate = str(Rate) + '%'
    G_Date = ws['k' + number].value
    I_Date = ws['l' + number].value
    E_Num = ws['h' + number].value
    People = ws['j' + number].value
    if Standard == '技术协议':
        Result = '技术'
    else:
        Result = '标准'
    Date = date_process(I_Date, 0, ['年', '月', '日'])
    context = dict(
        Place=Place,
        Number=Number,
        Name=Name,
        S_M=S_M,
        Supplier=Supplier,
        O_N=O_N,
        G_N=G_N,
        U=U,
        E_N=E_N,
        Standard=Standard,
        Rate=Rate,
        G_Date=G_Date,
        I_Date=I_Date,
        E_Num=E_Num,
        People=People,
        Result=Result,
        Date=Date
    )
    word.render(context)


# 以序号和报告编号命名并保存文件
# Parameter
# path:     (type: str)        文档保存路径
# word:  (type: DocxTemplate)   待保存的文档
# ws:     (type: worksheet)   待提取信息的工作表
# number:   (type: str)         等待提取的行
def save_file(path, word, ws, number):
    str_name = str(ws['a' + number].value) + '-' + ws['p' + number].value[8:-1] + '.docx'
    word.save(path + '\\' + str_name)

    try:
        word.save(path + '\\' + str_name)
    except FileNotFoundError:
        raise FileNotFoundError(r'保存路径不存在！')


def transform(model_path, folder_path, ws, start, end):
    i = 3
    while int(ws['a' + str(i)].value) <= end:
        if int(ws['a' + str(i)].value) >= start:
            word = open_model(model_path)
            write_information(word, ws, str(i))
            save_file(folder_path, word, ws, str(i))
        i += 1
        if ws['a' + str(i)].value is None:
            return


def main():
    ws = open_excel_worksheet_by_filename('副本报告清单2024.07.23.xlsx')
    transform('model.docx', r'E:\Files\Excel\Test_2\result', ws, 771, 826)


if __name__ == '__main__':
    main()
